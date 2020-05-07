#!/usr/bin/env Python

import requests
import pandas as pd
import io
from analyses.munging import get_county, get_centroid
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

DAYMET_URL = 'https://daymet.ornl.gov/single-pixel/api/data'



class Weather:

    ###
    def from_dataframe( self, wth_df ):
        #check keys
        keys = [ 'year', 'yday', 'prcp', 'srad', 'swe', 'tmax', 'tmin', 'vp' ]
        leap_years = [ yr for yr in range( 1980, 2020, 4 ) ]
        for key in keys:
            if key not in wth_df.columns:
                print( 'Imported weather data missing key "%"'.format( key ) )

        wth_df = wth_df.drop( columns = [ 'f1' ], axis = 1 )

        self.init_yr = wth_df[ 'year' ].min()
        self.end_yr = wth_df[ 'year' ].max()

        self.data = pd.DataFrame()
        self.data[ 'year' ] = wth_df[ 'year' ]
        self.data[ 'day' ] = wth_df[ 'yday' ]
        self.data[ 'dayL' ] = wth_df[ 'dayl' ]/3600
        self.data[ 'radn' ] = wth_df[ 'srad' ] * wth_df[ 'dayl' ] / 3600 * 0.0036
        self.data[ 'maxt' ] = wth_df[ 'tmax' ]
        self.data[ 'mint' ] = wth_df[ 'tmin' ]
        self.data[ 'prcp' ] = wth_df[ 'prcp' ]
        self.data[ 'swe' ] = wth_df[ 'swe' ]
        self.data[ 'vp' ] = wth_df[ 'vp' ] * 0.001
        self.data[ 'rain' ] = 0.0
        self.data[ 'snow' ] = 0.0

        # check for leap years
        for lp_yr in leap_years:
            lp_day = self.data.loc[ ( self.data[ 'year' ] == lp_yr ) &
                ( self.data[ 'day' ] == 365 ) ].copy( deep = True )
            lp_day[ 'day' ] = 366
            lp_day[ 'yday' ] = 366

            self.data = self.data.append( lp_day, ignore_index = True, sort = False )

        self.data = self.data.sort_values( by = [ 'year', 'day' ] )

        # check is snow-water equivalent increases next day
        for idx, row in self.data.iterrows():
            if idx == 0:
                self.data.loc[ idx:idx, 'snow' ] = 0.0
                self.data.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]
                continue
            elif idx == len( self.data ) - 1:
                self.data.loc[ idx:idx, 'snow' ] = 0.0
                self.data.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]
                continue
            else:
                cur = row[ 'swe' ]
                next = self.data.iloc[idx+1][ 'swe' ]
                if next > cur:
                    self.data.loc[ idx:idx, 'snow' ] = row[ 'prcp' ]
                    self.data.loc[ idx:idx, 'rain' ] = 0.0
                elif ( next > 0.0 ) & ( next == cur ):
                    self.data.loc[ idx:idx, 'snow' ] = row[ 'prcp' ]
                    self.data.loc[ idx:idx, 'rain' ] = 0.0
                else:
                    self.data.loc[ idx:idx, 'snow' ] = 0.0
                    self.data.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]

        self.data = self.data[[ 'year', 'day', 'radn', 'maxt', 'mint', 'rain',
            'snow', 'vp', 'dayL' ]]

        return self

    ###
    def from_daymet( self, lat, lon, startyr, endyr ):
        ### Daymet variables and units
        # day length (s/day)
        # min_temp (C)
        # max_temp (C)
        # precip (mm)
        # radiation (W/m2)
        # snow-water equiv. (kg/m2)
        # vapor pressure (Pa)

        attributes = [ 'dayl','prcp', 'srad','swe', 'tmax','tmin','vp' ]
        leap_years = [ yr for yr in range( 1980, 2020, 4 ) ]
        year_arr = [ str( startyr + i ) for i in range( endyr - startyr + 1 ) ]

        self.lat = lat
        self.lon = lon

        payload = {
            'lat': str( lat ),
            'lon': str( lon ),
            'vars': ','.join( attributes ),
            'years': ','.join( year_arr )
        }
        req = requests.get( DAYMET_URL, params = payload )
        wth_df = pd.read_csv( io.StringIO( req.text ), sep = ',', header = 6 )

        # day of year
        wth_df[ 'day' ] = wth_df[ 'yday' ]

        # daylength (hours)
        wth_df[ 'dayL' ] = round( wth_df[ 'dayl (s)' ]/3600, 1 )

        # solar radiation (MJ/m2)
        wth_df[ 'radn' ] = round( wth_df[ 'srad (W/m^2)' ] * wth_df[ 'dayl (s)' ] / 3600 * 0.0036, 1 )

        # max temperature (deg C)
        wth_df[ 'maxt' ] = round( wth_df[ 'tmax (deg c)' ], 1 )

        # min temperature (deg C)
        wth_df[ 'mint' ] = round( wth_df[ 'tmin (deg c)' ], 1 )

        # vapor pressure (kPa)
        wth_df[ 'vp' ] = round( wth_df[ 'vp (Pa)' ] * 0.001, 1 )

        # snow and rain (mm)
        wth_df[ 'rain' ] = 0.0
        wth_df[ 'snow' ] = 0.0

        # The Daymet calendar is based on a standard calendar year. All Daymet
        # years have 1 - 365 days, including leap years. For leap years, the Daymet
        # database includes leap day. Values for December 31 are discarded from
        # leap years to maintain a 365-day year.
        for lp_yr in leap_years:
            lp_day = wth_df.loc[ ( wth_df[ 'year' ] == lp_yr ) &
                ( wth_df[ 'day' ] == 365 ) ].copy( deep = True )
            lp_day[ 'day' ] = 366
            lp_day[ 'yday' ] = 366

            wth_df = wth_df.append( lp_day, ignore_index = True )

        wth_df = wth_df.sort_values( by = [ 'year', 'yday' ] )

        # check if snow-water equivalent increases next day
        for idx, row in wth_df.iterrows():
            if idx == 0:
                wth_df.iloc[idx][ 'snow' ] = 0.0
                wth_df.iloc[idx][ 'rain' ] = row[ 'prcp (mm/day)' ]
                continue
            elif idx == len( wth_df ) - 1:
                wth_df.iloc[idx][ 'snow' ] = 0.0
                wth_df.iloc[idx][ 'rain' ] = row[ 'prcp (mm/day)' ]
                continue
            else:
                cur = row[ 'swe (kg/m^2)' ]
                next = wth_df.iloc[idx+1][ 'swe (kg/m^2)' ]
                if next > cur:
                    wth_df.iloc[idx][ 'snow' ] = row[ 'prcp (mm/day)' ]
                    wth_df.iloc[idx][ 'rain' ] = 0.0
                elif ( next > 0.0 ) & ( next == cur ):
                    wth_df.iloc[idx][ 'snow' ] = row[ 'prcp (mm/day)' ]
                    wth_df.iloc[idx][ 'rain' ] = 0.0
                else:
                    wth_df.iloc[idx][ 'snow' ] = 0.0
                    wth_df.iloc[idx][ 'rain' ] = row[ 'prcp (mm/day)' ]

        wth_df = wth_df[ [ 'year', 'day', 'radn', 'maxt', 'mint', 'rain', 'snow',
            'vp', 'dayL' ] ]

        headers = ' '.join( [ 'year', 'day', 'radn', 'maxt', 'mint', 'rain',
            'snow', 'vp', 'dayL' ] )
        units = ' '.join( [ '()', '()', '(MJ/m^2)', '(oC)', '(oC)', '(mm)', '(mm)',
            '(kPa)', '(hours)' ] )

        self.data = wth_df

        return self

    ###
    def write_met_file( self, filepath ):
        # dump met file with Windows line endings
        if self.lat == None:
            lat = ''
            lon = ''
        else:
            lat = self.lat
            lon = self.lon

        if filepath:
            headers = ' '.join( [ 'year', 'day', 'radn', 'maxt', 'mint', 'rain',
                'snow', 'vp', 'dayL' ] )
            units = ' '.join( [ '()', '()', '(MJ/m^2)', '(oC)', '(oC)', '(mm)',
                '(mm)', '(kPa)', '(hours)' ] )

            metfile = open( filepath, 'w' )
            metfile.write( '[weather.met.weather]\r\n' )
            metfile.write( 'stateionname = Daymet weather\r\n')
            metfile.write( 'latitude = {} (DECIMAL DEGREES)\r\n'.format(
                lat ) )
            metfile.write( 'longitude = {} (DECIMAL DEGREES)\r\n'.format(
                lon ) )
            metfile.write( 'tav = ' +
                str( round( self.data[ 'maxt' ].mean(), 1 ) ) + '\r\n' )
            metfile.write( 'amp = ' +
                str( round( self.data[ 'maxt' ].max(), 1 ) ) + '\r\n' )
            metfile.write( '!Weather generated using ISU Foresite framework\r\n')
            metfile.write( headers + '\r\n' )
            metfile.write( units + '\r\n' )
            metfile.write( self.data.to_csv( sep = ' ', header = False,
                index = False, line_terminator='\r\n' ) )
            metfile.close()

    def add_daymet_spinup( self, lat, lon, init_yr, end_yr ):
        self.lat = lat
        self.lon = lon
        self.init_yr = init_yr

        attributes = [ 'weather_sample_id', 'dayl','prcp', 'srad','swe',
            'tmax','tmin','vp' ]
        leap_years = [ yr for yr in range( 1980, 2020, 4 ) ]

        # get spinup data from Daymet
        spup_start = init_yr
        spup_end = end_yr
        year_arr = [ str( init_yr + i ) for i in range( end_yr - init_yr + 1 ) ]
        payload = {
            'lat': lat,
            'lon': lon,
            'vars': ','.join( attributes ),
            'years': ','.join( year_arr )
        }
        req = requests.get( DAYMET_URL, params = payload )
        spinup_df = pd.read_csv( io.StringIO( req.text ), sep = ',', header = 6 )

        wth_df = pd.DataFrame()
        wth_df[ 'year' ] = spinup_df[ 'year' ]
        wth_df[ 'day' ] = spinup_df[ 'yday' ]
        wth_df[ 'dayL' ] = spinup_df[ 'dayl (s)' ]/3600
        wth_df[ 'radn' ] = spinup_df[ 'srad (W/m^2)' ] * spinup_df[ 'dayl (s)' ] / 3600 * 0.0036
        wth_df[ 'maxt' ] = spinup_df[ 'tmax (deg c)' ]
        wth_df[ 'mint' ] = spinup_df[ 'tmin (deg c)' ]
        wth_df[ 'prcp' ] = spinup_df[ 'prcp (mm/day)' ]
        wth_df[ 'swe' ] = spinup_df[ 'swe (kg/m^2)' ]
        wth_df[ 'vp' ] = spinup_df[ 'vp (Pa)' ] * 0.001
        wth_df[ 'rain' ] = 0.0
        wth_df[ 'snow' ] = 0.0

        # check for leap years
        for lp_yr in leap_years:
            lp_day = wth_df.loc[ ( wth_df[ 'year' ] == lp_yr ) &
                ( wth_df[ 'day' ] == 365 ) ].copy( deep = True )
            lp_day[ 'day' ] = 366
            lp_day[ 'yday' ] = 366

            wth_df = wth_df.append( lp_day, ignore_index = True, sort = False )

        wth_df = wth_df.sort_values( by = [ 'year', 'day' ] )

        # check is snow-water equivalent increases next day
        for idx, row in wth_df.iterrows():
            if idx == 0:
                wth_df.loc[ idx:idx, 'snow' ] = 0.0
                wth_df.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]
                continue
            elif idx == len( wth_df ) - 1:
                wth_df.loc[ idx:idx, 'snow' ] = 0.0
                wth_df.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]
                continue
            else:
                cur = row[ 'swe' ]
                next = wth_df.iloc[idx+1][ 'swe' ]
                if next > cur:
                    wth_df.loc[ idx:idx, 'snow' ] = row[ 'prcp' ]
                    wth_df.loc[ idx:idx, 'rain' ] = 0.0
                elif ( next > 0.0 ) & ( next == cur ):
                    wth_df.loc[ idx:idx, 'snow' ] = row[ 'prcp' ]
                    wth_df.loc[ idx:idx, 'rain' ] = 0.0
                else:
                    wth_df.loc[ idx:idx, 'snow' ] = 0.0
                    wth_df.loc[ idx:idx, 'rain' ] = row[ 'prcp' ]

        wth_df = wth_df[[ 'year', 'day', 'radn', 'maxt', 'mint', 'rain',
            'snow', 'vp', 'dayL' ]]

        self.data = wth_df.append( self.data, sort = False )
        self.data = self.data.round( 2 )

def create_excel_met(lat, long, start_year, end_year, county_name):
    wth_obj = Weather().from_daymet(lat, long, 1980, 2019)
    wth_df = wth_obj.data
    tav = round(wth_df[ 'maxt' ].mean(), 1)
    amp = round(wth_df['maxt'].max(), 1)
    #greene_df.to_excel('greene.xlsx', index=False)
    wb = Workbook()
    ws = wb.active
    ws.alignment = Alignment(horizontal="left")
    for r in dataframe_to_rows(wth_df, index=False, header=True):
        ws.append(r)
    for cells in ws.iter_rows():
        for cell in cells:
            cell.alignment = Alignment(horizontal="left")
    ws.insert_rows(2)
    ws['A2'] = '()'
    ws['B2'] = '()'
    ws['C2'] = '(MJ/m2)'
    ws['D2'] = '(oC)'
    ws['E2'] = '(oC)'
    ws['F2'] = '(mm)'
    ws['G2'] = '(mm)'
    ws['H2'] = '(kPa)'
    ws['I2'] = '(hours)'
    ws.insert_rows(1)
    ws['A1'] = '!Weather generated using C-CHANGE Foresite framework'
    ws.insert_rows(1)
    ws['A1'] = f'amp = {amp}'
    ws.insert_rows(1)
    ws['A1'] = f'tav = {tav}'
    ws.insert_rows(1)
    ws['A1'] = f'longitude = {long} (DECIMAL DEGREES)'
    ws.insert_rows(1)
    ws['A1'] = f'latitude = {lat} (DECIMAL DEGREES)'
    ws.insert_rows(1)
    ws['A1'] = 'stationname = Daymet weather'
    ws.insert_rows(1)
    ws['A1'] = '[weather.met.weather]'
    wb.save(f'apsim_files/{county_name}/met_files/{county_name}.xlsx')


'''
Get the weather for given centroid and write to a .met file

Args:
    lat {float} -- latitude of centroid
    long {float} -- longitude of centroid
    year_star {int} -- starting year of weather data
    year_end {int} -- ending year of weather data
    path {str} -- path to write the met files
    filename {str} -- name to give the .met file

Returns:
    None
'''
def create_met(lat, long, start_year, end_year, filename, path='apsim_files/met_files'):
    weather_obj = Weather().from_daymet(lat, long, start_year, end_year)
    weather_obj.write_met_file(f'{path}/{filename}.met')


    """
Create all met files for each county in a given geometry.

Args:
    dbconn {database connection} -- connection to postgresql database
    counties {array/list} -- every county to be run on in the geometry
    table {str} -- name of the table within the database to get geometries from
    id_column {str} -- table column that has unique ids (in this case fips) for each county.
    geo_col {str} -- table column that has geometry.
    name_col {str} -- column that has the name for each county - can just use fips columns if no county names in table.

Returns:
    Met file for each county in counties.
"""
def create_all_met(dbconn, counties, table, id_col='fips', geo_col='wkb_geometry', name_col='county'):
    for i in counties:
        county = get_county(dbconn, table, i, geo_col)
        county_name = county[name_col][0].replace(" ", "_")
        print(f'Geopandas table for {county_name} county created.')
        centroid = get_centroid(county, id_col, geo_col)
        print(f'Centroid located at {centroid}.')
        weather = create_excel_met(centroid[0], centroid[1], 1980, 2019, county_name)
        print(f"Met file for {county_name}/{i} at location {centroid} created.")

def create_all_excel_met(dbconn, counties, table, id_col='fips', geo_col='wkb_geometry', name_col='county'):
    for i in counties:
        county = get_county(dbconn, table, i, geo_col)
        county_name = county[name_col][0].replace(" ", "_")
        print(f'Geopandas table for {county_name} county created.')
        centroid = get_centroid(county, id_col, geo_col)
        print(f'Centroid located at {centroid}.')
        if not os.path.exists(f'apsim_files/{county_name}/met_files'):
            os.makedirs(f'apsim_files/{county_name}/met_files')
        create_excel_met(centroid[0], centroid[1], 1980, 2019, county_name)
        print(f"Met file for {county_name}/{i} at location {centroid} created.")